"""Report export endpoints.

Renders the executive PDF, technical PDF, findings CSV or JSON for a completed
analysis. PDFs are streamed back to the caller; large fleets typically request
these asynchronously and receive a signed storage URL, but the synchronous path
is provided for convenience and for the single-device case.
"""

from __future__ import annotations

import io
import tempfile

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy.orm import Session

from ..database import get_db
from ..models import User, Analysis, AnalysisStatus, Organization, Finding
from ..security import current_user

from firewallguard.report import generator as gen

router = APIRouter(prefix="/api/v1", tags=["reports"])


def _load_complete(analysis_id: str, user: User, db: Session) -> Analysis:
    analysis = db.get(Analysis, analysis_id)
    if analysis is None or analysis.organization_id != user.organization_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Analysis not found")
    if analysis.status != AnalysisStatus.complete:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT,
                            detail=f"Analysis status is {analysis.status.value}, not complete")
    return analysis


def _branding(db: Session, organization_id: str) -> dict | None:
    """Assemble white-label branding for the org, or None for default branding."""
    org = db.get(Organization, organization_id)
    if org is None or not (org.brand_company_name or org.brand_logo_url
                           or org.brand_primary_color or org.brand_contact):
        return None
    return {"company_name": org.brand_company_name, "logo_url": org.brand_logo_url,
            "primary_color": org.brand_primary_color, "contact": org.brand_contact}


def _pdf_stream(builder, analysis_dict, filename: str, branding: dict | None) -> StreamingResponse:
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        builder(analysis_dict, tmp.name, branding)
        tmp.seek(0)
        data = open(tmp.name, "rb").read()
    return StreamingResponse(io.BytesIO(data), media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/analyses/{analysis_id}/report/executive")
def executive_report(analysis_id: str,
                     user: User = Depends(current_user),
                     db: Session = Depends(get_db)) -> StreamingResponse:
    analysis = _load_complete(analysis_id, user, db)
    return _pdf_stream(gen.build_executive_pdf, analysis.result_json,
                       f"firewallguard-executive-{analysis_id[:8]}.pdf",
                       _branding(db, user.organization_id))


@router.get("/analyses/{analysis_id}/report/technical")
def technical_report(analysis_id: str,
                     user: User = Depends(current_user),
                     db: Session = Depends(get_db)) -> StreamingResponse:
    from sqlalchemy import select
    analysis = _load_complete(analysis_id, user, db)
    result = analysis.result_json or {}

    # Enrich findings with live statuses from the findings table
    device_id = analysis.device_id
    live_findings = db.scalars(
        select(Finding).where(
            Finding.organization_id == user.organization_id,
            Finding.device_id == device_id,
        )
    ).all()

    # Build fingerprint -> live-finding map
    status_map: dict[str, dict] = {}
    for lf in live_findings:
        fp = f"{lf.rule_id}::{lf.object_type}::{lf.object_name}"
        status_map[fp] = {
            "status": lf.status.value,
            "status_label": lf.status.value.replace("_", " ").title(),
            "first_seen": lf.first_seen_at.isoformat() if lf.first_seen_at else None,
            "last_seen": lf.last_seen_at.isoformat() if lf.last_seen_at else None,
        }

    # Enrich each snapshot finding
    enriched_findings = []
    for f in result.get("findings", []):
        fp = f"{f.get('rule_id','')}::{f.get('object_type','')}::{f.get('object_name','')}"
        live = status_map.get(fp, {})
        enriched_findings.append({
            **f,
            "status": live.get("status", "open"),
            "status_label": live.get("status_label", "Open"),
            "first_seen": live.get("first_seen"),
            "last_seen": live.get("last_seen"),
        })

    enriched = dict(result)
    enriched["findings"] = enriched_findings
    enriched["device_id"] = device_id

    return _pdf_stream(gen.build_technical_pdf, enriched,
                       f"firewallguard-technical-{analysis_id[:8]}.pdf",
                       _branding(db, user.organization_id))


@router.get("/analyses/{analysis_id}/export/xlsx")
def export_xlsx(analysis_id: str,
                user: User = Depends(current_user),
                db: Session = Depends(get_db)) -> StreamingResponse:
    """Export the analysis findings as a formatted Excel workbook."""
    analysis = _load_complete(analysis_id, user, db)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    result = analysis.result_json or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Findings"

    # Summary header.
    dev = result.get("device", {})
    score = result.get("score", {})
    ws.append([f"FirewallGuard AI — {dev.get('model', 'Device')} {dev.get('serial', '')}"])
    ws.append([f"Score {score.get('score', 0)}/100  Grade {score.get('grade', '')}",
               f"Findings: {result.get('finding_count', 0)}"])
    ws.append([])

    headers = ["Severity", "Rule", "Category", "Title", "Object Type", "Object",
               "Exploitability", "Remediation"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="1D4ED8")
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    sev_color = {"Critical": "FF4D4D", "High": "FF8A3D", "Medium": "F5C451",
                 "Low": "4A9EFF", "Info": "7A879B"}
    for f in result.get("findings", []):
        ws.append([f.get("severity", ""), f.get("rule_id", ""), f.get("category", ""),
                   f.get("title", ""), f.get("object_type", ""), f.get("object_name", ""),
                   f.get("exploitability", ""), f.get("remediation", "")])
        c = ws.cell(row=ws.max_row, column=1)
        c.font = Font(bold=True, color=sev_color.get(f.get("severity", ""), "000000"))

    widths = [12, 14, 20, 46, 16, 26, 14, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="findings-{analysis_id[:8]}.xlsx"'})


@router.get("/devices/{device_id}/compare/report")
def comparison_report(device_id: str,
                      previous: str, current: str,
                      user: User = Depends(current_user),
                      db: Session = Depends(get_db)) -> StreamingResponse:
    """Download a PDF comparison report between two TSRs for a device."""
    from ..models import Device, Analysis as AnalysisModel
    device = db.get(Device, device_id)
    if device is None or device.organization_id != user.organization_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Device not found")

    prev = db.get(AnalysisModel, previous)
    curr = db.get(AnalysisModel, current)
    if prev is None or prev.organization_id != user.organization_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Previous analysis not found")
    if curr is None or curr.organization_id != user.organization_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Current analysis not found")
    if prev.status != AnalysisStatus.complete or curr.status != AnalysisStatus.complete:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Both analyses must be complete")

    branding = _branding(db, user.organization_id)
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        gen.build_comparison_pdf(prev.result_json or {}, curr.result_json or {},
                                 tmp.name, branding)
        tmp.seek(0)
        data = open(tmp.name, "rb").read()
    return StreamingResponse(io.BytesIO(data), media_type="application/pdf",
                             headers={"Content-Disposition":
                                      f'attachment; filename="comparison-{device_id[:8]}.pdf"'})


@router.get("/analyses/{analysis_id}/export/csv")
def export_csv(analysis_id: str,
               user: User = Depends(current_user),
               db: Session = Depends(get_db)) -> StreamingResponse:
    analysis = _load_complete(analysis_id, user, db)
    csv_text = gen.to_csv_string(analysis.result_json)
    return StreamingResponse(io.StringIO(csv_text), media_type="text/csv",
                             headers={"Content-Disposition":
                                      f'attachment; filename="findings-{analysis_id[:8]}.csv"'})


@router.get("/analyses/{analysis_id}/export/json")
def export_json(analysis_id: str,
                user: User = Depends(current_user),
                db: Session = Depends(get_db)) -> JSONResponse:
    analysis = _load_complete(analysis_id, user, db)
    payload = {k: v for k, v in analysis.result_json.items() if k != "snapshot"}
    return JSONResponse(content=payload)
